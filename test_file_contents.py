import zipfile, csv
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from import_os_path import ZIP_DIR


def test_pdf_content():
    with zipfile.ZipFile(ZIP_DIR) as zfile:
        with zfile.open('sample_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            assert "Fun fun fun" in reader.pages[0].extract_text()
            assert len(reader.pages) == 1


def test_xlsx_content():
    with zipfile.ZipFile(ZIP_DIR) as zip_files:
        with zip_files.open('sample_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=3, column=2).value == 'Mara'


def test_csv_content():
    with zipfile.ZipFile(ZIP_DIR) as zfile:
        with zfile.open("sample_csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            csv_reader = list(csv.reader(content.splitlines(), delimiter=','))
            first_row_values = csv_reader[0]
            assert first_row_values[0] == '1'
            assert first_row_values[1] == 'a'
            assert len(csv_reader) == 5
